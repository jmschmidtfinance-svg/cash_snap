#!/usr/bin/env python3
"""
Daily cash snap pipeline (single subsidiary) -- v2.

Cash perimeter = an explicit allowlist of account internal IDs (CASH_ACCOUNT_IDS; default
First Bank 223 + Undeposited Funds 122). "Cash" is the combined balance; a Deposit (UF -> Bank)
is internal and nets to zero.

Reconciliation is a stateful ROLL-FORWARD:
  last *reported* balance (from state file)  +  captured movements  ==  today's computed balance
Captured movements = postings to cash accounts that are dated in the window OR were created
since the last snapshot (catches back-dated entries). State is persisted to state/cash_state.json
and committed back to the repo by the workflow, which also serves as an audit ledger.

Design rule: CODE COMPUTES, THE LLM EXPLAINS. The model only writes the CFO narrative and the
CEO text-message block from finished numbers.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import json
import os
import smtplib
import sys
from collections import defaultdict
from email.message import EmailMessage

import requests
from requests_oauthlib import OAuth1

# --------------------------------------------------------------------------------------
# Config (env vars only)
# --------------------------------------------------------------------------------------
NS_ACCOUNT_ID   = os.environ["NS_ACCOUNT_ID"]
NS_CONSUMER_KEY = os.environ["NS_CONSUMER_KEY"]
NS_CONSUMER_SEC = os.environ["NS_CONSUMER_SECRET"]
NS_TOKEN_ID     = os.environ["NS_TOKEN_ID"]
NS_TOKEN_SEC    = os.environ["NS_TOKEN_SECRET"]

ANTHROPIC_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-6")
KB_PATH         = os.environ.get("KB_PATH", "netsuite_knowledge_base.md")
STATE_PATH      = os.environ.get("STATE_PATH", "state/cash_state.json")

SMTP_HOST = os.environ["SMTP_HOST"]
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ["SMTP_USER"]
SMTP_PASS = os.environ["SMTP_PASSWORD"]
MAIL_FROM = os.environ.get("MAIL_FROM", SMTP_USER)
MAIL_TO   = os.environ["MAIL_TO"]

THRESHOLD = float(os.environ.get("SNAP_THRESHOLD") or "50000")
REPORT_DATE_OVERRIDE = os.environ.get("REPORT_DATE")

# Sign convention VERIFIED 2026-06-29: cash inflow posts positive. Reconciliation catches errors.
INFLOW_IS_POSITIVE = True

# The cash perimeter is an explicit ALLOWLIST of account internal IDs -- not an account-type
# filter. We track only the operating account and Undeposited Funds:
#     223 = First Bank (operating)        122 = Undeposited Funds
# Selecting by accttype='Bank' also swept in Brokerage (312), First Bank of the Lake (311) and
# Petty Cash (224) -- balances that aren't part of the operating-cash view and that added noise to
# the daily net change. An explicit id list is the only clean selector (UF isn't even type 'Bank';
# it is 'OthCurrAsset'). Override via CASH_ACCOUNT_IDS (comma-separated internal ids) when an
# account is opened/closed. NOTE: changing this set re-bases total cash, so the next run should
# re-bootstrap the prior baseline (delete state/cash_state.json) to avoid a one-time roll-forward
# mismatch equal to the dropped accounts' balances.
CASH_ACCOUNT_IDS = tuple(
    int(x) for x in (os.environ.get("CASH_ACCOUNT_IDS") or "223,122").replace(" ", "").split(",") if x
)

PAYROLL_ACCOUNT_IDS: set[int] = set()
TOP_CEO_ITEMS = 8            # how many vendors/customers to itemize before "All others"
EPSILON = 0.01

# Project (called "job" internally in NetSuite -- Jobs were renamed Projects in the UI but the
# SuiteQL record kept the name). The GL project on an invoice line is the RABB-IT column
# custcol_r_it_reporting_project (populated on Production invoices; null on Service, which is fine
# -- those fall through to the Service bucket). The project record's Production/Service flag is the
# custom field custentity_r_it_class, matching the Class master: 1 = Production, 2 = Service.
PROJECT_GL_FIELD    = "custcol_r_it_reporting_project"
PROJECT_CLASS_FIELD = "custentity_r_it_class"
CLASS_PRODUCTION    = 1
CLASS_SERVICE       = 2


# --------------------------------------------------------------------------------------
# SuiteQL client
# --------------------------------------------------------------------------------------
def _rest_host() -> str:
    return f"{NS_ACCOUNT_ID.lower().replace('_', '-')}.suitetalk.api.netsuite.com"


def _auth() -> OAuth1:
    return OAuth1(
        client_key=NS_CONSUMER_KEY, client_secret=NS_CONSUMER_SEC,
        resource_owner_key=NS_TOKEN_ID, resource_owner_secret=NS_TOKEN_SEC,
        signature_method="HMAC-SHA256", realm=NS_ACCOUNT_ID,
    )


def run_suiteql(sql: str, page_size: int = 1000) -> list[dict]:
    url = f"https://{_rest_host()}/services/rest/query/v1/suiteql"
    headers = {"Content-Type": "application/json", "Prefer": "transient"}
    auth = _auth()
    rows, offset = [], 0
    while True:
        resp = requests.post(url, params={"limit": page_size, "offset": offset},
                             headers=headers, auth=auth, json={"q": sql}, timeout=60)
        if resp.status_code != 200:
            raise RuntimeError(f"SuiteQL {resp.status_code}: {resp.text[:500]}")
        body = resp.json()
        rows.extend(body.get("items", []))
        if not body.get("hasMore"):
            return rows
        offset += page_size


def _types_in(types: tuple[str, ...]) -> str:
    return ", ".join(f"'{t}'" for t in types)


def _acct_selector(acct_types: tuple[str, ...], extra_ids: tuple[int, ...] = ()) -> str:
    """SQL predicate: account is one of these types OR one of these explicit internal IDs.
    With acct_types empty it becomes a pure id allowlist (the cash perimeter); with extra_ids
    empty it is a pure type filter (the AR/AP GL totals)."""
    parts = []
    if acct_types:
        parts.append(f"a.accttype IN ({_types_in(acct_types)})")
    if extra_ids:
        parts.append(f"a.id IN ({', '.join(str(int(i)) for i in extra_ids)})")
    if not parts:
        return "1=1"
    return "(" + " OR ".join(parts) + ")"


def _class_int(x):
    """Coerce a project class id to int. SuiteQL hands back the custom-field value as a STRING
    ('1'/'2') over REST, so a raw `== 1` int comparison silently fails and every job collapses to
    Service. Returns None when the value is absent or unparseable."""
    try:
        return int(x)
    except (TypeError, ValueError):
        return None


def _parse_ns_date(s):
    """Parse a NetSuite display date ('M/D/YYYY', or an ISO 'YYYY-MM-DD') to a date; None if it
    can't be parsed. Used only for derived fields like days-overdue -- never for query bounds."""
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(str(s)[:10] if fmt == "%Y-%m-%d" else str(s), fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------------------
# Dates  (fire after midnight Eastern; report the day that just closed)
# --------------------------------------------------------------------------------------
def resolve_report_date() -> dt.date:
    if REPORT_DATE_OVERRIDE:
        return dt.date.fromisoformat(REPORT_DATE_OVERRIDE)
    return (dt.datetime.utcnow() - dt.timedelta(days=1)).date()


def previous_business_day(d: dt.date) -> dt.date:
    d -= dt.timedelta(days=1)
    while d.weekday() >= 5:
        d -= dt.timedelta(days=1)
    return d


# --------------------------------------------------------------------------------------
# Queries
# --------------------------------------------------------------------------------------
def balances_as_of(d: dt.date, acct_types: tuple[str, ...],
                   created_on_or_before: dt.date | None = None,
                   extra_account_ids: tuple[int, ...] = ()) -> list[dict]:
    """Posting balance as of trandate <= d.

    If created_on_or_before is given, also require the entry to have been CREATED by then --
    i.e. reconstruct the balance as it actually stood at close, excluding anything back-posted
    afterward. Used to synthesize a clean prior baseline on the bootstrap run.
    extra_account_ids folds in non-type-matched cash accounts (Undeposited Funds).
    """
    created_clause = (
        f"AND TRUNC(t.createddate) <= TO_DATE('{created_on_or_before.isoformat()}', 'YYYY-MM-DD')"
        if created_on_or_before else ""
    )
    sql = f"""
        SELECT a.id AS account_id, a.acctnumber, a.fullname, a.accttype,
               SUM(tal.amount) AS balance
        FROM   transactionaccountingline tal
        JOIN   transaction t ON t.id = tal.transaction
        JOIN   account a     ON a.id = tal.account
        WHERE  {_acct_selector(acct_types, extra_account_ids)}
          AND  tal.posting = 'T'
          AND  t.trandate <= TO_DATE('{d.isoformat()}', 'YYYY-MM-DD')
          {created_clause}
        GROUP BY a.id, a.acctnumber, a.fullname, a.accttype
    """
    return run_suiteql(sql)


def total_balance(d: dt.date, acct_types: tuple[str, ...],
                  created_on_or_before: dt.date | None = None,
                  extra_account_ids: tuple[int, ...] = ()) -> float:
    return sum(float(r["balance"])
               for r in balances_as_of(d, acct_types, created_on_or_before, extra_account_ids))


def cash_movements(prior: dt.date, report: dt.date, include_created: bool) -> list[dict]:
    """Cash-account postings new since the prior snapshot.

    include_created=True: dated in window OR created since prior_date. Used by both the logged
      path (prior = last reported balance) and the bootstrap path (prior = reconstructed
      created-on-or-before baseline) -- both are true roll-forwards that surface back-posts.
    include_created=False: dated in window only. Kept for a naive recompute if ever needed.
    """
    created_clause = (
        f"OR TRUNC(t.createddate) > TO_DATE('{prior.isoformat()}', 'YYYY-MM-DD')"
        if include_created else ""
    )
    sql = f"""
        SELECT t.id AS tran_id, t.tranid, t.trandate, t.createddate,
               t.type AS type_code, BUILTIN.DF(t.type) AS type_label,
               a.id AS account_id, a.acctnumber, a.fullname AS account_name, a.accttype,
               BUILTIN.DF(t.entity) AS entity_name, t.memo, tal.amount
        FROM   transactionaccountingline tal
        JOIN   transaction t ON t.id = tal.transaction
        JOIN   account a     ON a.id = tal.account
        WHERE  {_acct_selector((), CASH_ACCOUNT_IDS)}
          AND  tal.posting = 'T'
          AND  t.trandate <= TO_DATE('{report.isoformat()}', 'YYYY-MM-DD')
          AND  ( t.trandate > TO_DATE('{prior.isoformat()}', 'YYYY-MM-DD') {created_clause} )
        ORDER BY tal.amount
    """
    return run_suiteql(sql)


def cash_in_by_project(prior: dt.date, report: dt.date, include_created: bool) -> list[dict]:
    """Customer payments in the roll-forward window, split to the project they collected against.

    Each payment -> applied invoice(s) via PreviousTransactionLineLink (foreignamount = the amount
    applied to that invoice); each invoice -> its GL project (the line-level reporting-project field,
    taken as one project per invoice) -> the project record + its Production/Service class. Same
    window as cash_movements, so the attributed total ties to the AR-collections inflow. One row per
    (payment, invoice) link; a payment spanning multiple invoices/projects yields multiple rows."""
    created_clause = (
        f"OR TRUNC(pay.createddate) > TO_DATE('{prior.isoformat()}', 'YYYY-MM-DD')"
        if include_created else ""
    )
    sql = f"""
        SELECT pay.tranid         AS payment_no,
               pay.trandate       AS payment_date,
               BUILTIN.DF(pay.entity) AS customer,
               inv.tranid         AS invoice_no,
               link.foreignamount AS applied_amt,
               proj.entityid      AS project_num,
               proj.companyname   AS project_name,
               proj.{PROJECT_CLASS_FIELD} AS class_id
        FROM   transaction pay
        JOIN   PreviousTransactionLineLink link
                 ON link.nextdoc = pay.id AND link.previoustype = 'CustInvc'
        JOIN   transaction inv ON inv.id = link.previousdoc
        LEFT JOIN job proj ON proj.id = (
                 SELECT MIN(tl.{PROJECT_GL_FIELD}) FROM transactionline tl
                 WHERE  tl.transaction = inv.id AND tl.{PROJECT_GL_FIELD} IS NOT NULL)
        WHERE  pay.type = 'CustPymt'
          AND  pay.trandate <= TO_DATE('{report.isoformat()}', 'YYYY-MM-DD')
          AND  ( pay.trandate > TO_DATE('{prior.isoformat()}', 'YYYY-MM-DD') {created_clause} )
    """
    rows = run_suiteql(sql)
    for r in rows:
        r["class_id"] = _class_int(r.get("class_id"))
    return rows


def overdue_bills(report: dt.date) -> dict:
    """Open vendor bills whose due date is on or before the report date, at remaining (unpaid)
    balance. The boundary is INCLUSIVE (<=): a bill due ON the report date is past due by the time
    the snap is read the next morning, and a strict `<` dropped whole days of subcontractor bills
    (e.g. all of a day's roofing subs) from the total -- the source of the ~$70k-vs-actual gap.
    Returns the aggregate PLUS the full bill list so the summary and the overdue workbook share a
    single query."""
    sql = f"""
        SELECT t.tranid            AS bill_no,
               BUILTIN.DF(t.entity) AS vendor,
               t.trandate,
               t.duedate,
               ABS(t.foreigntotal)  AS bill_amount,
               t.foreignamountunpaid AS unpaid,
               t.status
        FROM   transaction t
        WHERE  t.type = 'VendBill'
          AND  t.duedate <= TO_DATE('{report.isoformat()}', 'YYYY-MM-DD')
          AND  t.foreignamountunpaid > 0
        ORDER BY t.duedate, vendor
    """
    bills = []
    for r in run_suiteql(sql):
        due = _parse_ns_date(r.get("duedate"))
        bills.append({
            "bill_no": r.get("bill_no"), "vendor": r.get("vendor"),
            "trandate": r.get("trandate"), "duedate": r.get("duedate"),
            "days_overdue": (report - due).days if due else None,
            "bill_amount": round(float(r.get("bill_amount") or 0), 2),
            "unpaid": round(float(r.get("unpaid") or 0), 2),
            "status": r.get("status"),
        })
    total = round(sum(b["unpaid"] for b in bills), 2)
    return {"basis": "due on or before report date", "asof": report.isoformat(),
            "count": len(bills), "total": total, "bills": bills}


def _window_clause(alias: str, prior: dt.date, report: dt.date, include_created: bool) -> str:
    """trandate in (prior, report]  OR  created after prior (to catch back-posts) -- mirrors
    cash_movements so an attribution query sees exactly the same transactions."""
    created = (f"OR TRUNC({alias}.createddate) > TO_DATE('{prior.isoformat()}', 'YYYY-MM-DD')"
               if include_created else "")
    return (f"{alias}.trandate <= TO_DATE('{report.isoformat()}', 'YYYY-MM-DD') "
            f"AND ({alias}.trandate > TO_DATE('{prior.isoformat()}', 'YYYY-MM-DD') {created})")


def ap_paid_by_project(prior: dt.date, report: dt.date, include_created: bool) -> list[dict]:
    """AP cash paid out in the window, attributed to projects via the GL reporting project.

    Two disjoint paths, unioned:
      (A) Bill payments (VendPymt) -> applied bill via PreviousTransactionLineLink (foreignamount =
          amount paid on that bill) -> the bill's line-level reporting project.
      (B) Direct checks (Check) -> the check's own non-mainline (expense) lines, which carry the
          reporting project directly; amount = the line's debit.
    Same window as cash_movements, so the attributed total ties to the AP-disbursements bucket.
    One row per (payment, bill) for A and per (check, expense line) for B; amounts are positive."""
    sql = f"""
        SELECT payment_no, payment_date, vendor, source, reference, amt,
               project_num, project_name, class_id
        FROM (
            SELECT pay.tranid       AS payment_no,
                   pay.trandate     AS payment_date,
                   BUILTIN.DF(pay.entity) AS vendor,
                   'billpay'        AS source,
                   bill.tranid      AS reference,
                   link.foreignamount AS amt,
                   proj.entityid    AS project_num,
                   proj.companyname AS project_name,
                   proj.{PROJECT_CLASS_FIELD} AS class_id
            FROM   transaction pay
            JOIN   PreviousTransactionLineLink link
                     ON link.nextdoc = pay.id AND link.previoustype = 'VendBill'
            JOIN   transaction bill ON bill.id = link.previousdoc
            LEFT JOIN job proj ON proj.id = (
                     SELECT MIN(tl.{PROJECT_GL_FIELD}) FROM transactionline tl
                     WHERE  tl.transaction = bill.id AND tl.{PROJECT_GL_FIELD} IS NOT NULL)
            WHERE  pay.type = 'VendPymt' AND {_window_clause('pay', prior, report, include_created)}
            UNION ALL
            SELECT t.tranid          AS payment_no,
                   t.trandate        AS payment_date,
                   BUILTIN.DF(t.entity) AS vendor,
                   'check'           AS source,
                   BUILTIN.DF(tl.account) AS reference,
                   tl.amount         AS amt,
                   proj.entityid     AS project_num,
                   proj.companyname  AS project_name,
                   proj.{PROJECT_CLASS_FIELD} AS class_id
            FROM   transactionline tl
            JOIN   transaction t ON t.id = tl.transaction
            LEFT JOIN job proj ON proj.id = tl.{PROJECT_GL_FIELD}
            WHERE  t.type = 'Check' AND tl.mainline = 'F'
              AND  {_window_clause('t', prior, report, include_created)}
        )
    """
    rows = run_suiteql(sql)
    for r in rows:
        r["class_id"] = _class_int(r.get("class_id"))
    return rows


# Liability account types whose appearance in a journal's offsets signals debt principal.
_LIABILITY_TYPES = ("LongTermLiab", "OthCurrLiab", "OthLiab", "CredCard")


def journal_offsets(movements: list[dict]) -> dict:
    """For every Journal that hit the cash perimeter in this window, return its offsetting GL
    accounts (netted per account, dropping the cash-perimeter legs and any net-zero clearing
    legs) so the narrative model can state each journal's PURPOSE instead of leaving it in
    'Other / unclassified'. Grouping by account collapses paired intercompany-clearing legs to
    zero automatically. When the surviving offsets are a single liability account plus an interest
    line, we also hand over a deterministic suggested label (debt principal & interest)."""
    jids = sorted({r.get("tran_id") for r in movements
                   if (r.get("type_code") or "") == "Journal" and r.get("tran_id")})
    if not jids:
        return {}
    id_list = ", ".join(str(int(i)) for i in jids)
    cash_ids = ", ".join(str(int(i)) for i in CASH_ACCOUNT_IDS)
    sql = f"""
        SELECT t.tranid AS journal_no, acc.fullname AS account, acc.accttype,
               SUM(tal.amount) AS net_amt
        FROM   transaction t
        JOIN   transactionaccountingline tal ON tal.transaction = t.id
        JOIN   account acc ON acc.id = tal.account
        WHERE  t.id IN ({id_list}) AND tal.posting = 'T'
          AND  tal.account NOT IN ({cash_ids})
        GROUP BY t.tranid, acc.fullname, acc.accttype
        HAVING ABS(SUM(tal.amount)) > {EPSILON}
        ORDER BY t.tranid, net_amt
    """
    out: dict[str, dict] = {}
    for r in run_suiteql(sql):
        out.setdefault(r.get("journal_no"), {"offsets": []})["offsets"].append({
            "account": r.get("account"), "accttype": r.get("accttype"),
            "amount": round(float(r.get("net_amt") or 0), 2),
        })
    for d in out.values():
        offs = d["offsets"]
        liab = [o for o in offs if o["accttype"] in _LIABILITY_TYPES]
        interest = [o for o in offs if "interest" in (o["account"] or "").lower()]
        if len(liab) == 1 and interest:
            d["suggested_purpose"] = f"Debt Payment - {liab[0]['account']} (P&I)"
    return out
def load_state() -> dict | None:
    try:
        with open(STATE_PATH, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def save_state(report: dt.date, total_cash: float, balances: list[dict]) -> None:
    os.makedirs(os.path.dirname(STATE_PATH) or ".", exist_ok=True)
    payload = {
        "report_date": report.isoformat(),
        "total_cash": round(total_cash, 2),
        "balances": {str(r["account_id"]): {"name": r["fullname"],
                                             "balance": round(float(r["balance"]), 2)}
                     for r in balances},
    }
    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# --------------------------------------------------------------------------------------
# Bucketizing
# --------------------------------------------------------------------------------------
def signed(amount) -> float:
    a = float(amount)
    return a if INFLOW_IS_POSITIVE else -a


def classify(row: dict) -> str:
    tc = (row.get("type_code") or "").strip()
    if tc in ("Transfer", "Deposit"):
        return "Internal transfer"          # Deposit = UF -> Bank, internal to cash perimeter
    if tc == "CustPymt":
        return "AR collections"
    if tc == "Paycheck" or row.get("account_id") in PAYROLL_ACCOUNT_IDS:
        return "Payroll"
    if tc in ("VendPymt", "Check"):
        return "AP disbursements"
    return "Other / unclassified"           # Journals etc. -- see KB sec.7 / sec.10


# --------------------------------------------------------------------------------------
# Summary assembly
# --------------------------------------------------------------------------------------
def _itemize(rows: list[dict], inflow: bool) -> list[dict]:
    """Group external (non-internal) movements by entity, signed, sorted by magnitude."""
    agg: dict[str, float] = defaultdict(float)
    for r in rows:
        if classify(r) == "Internal transfer":
            continue
        amt = signed(r["amount"])
        if (amt > 0) != inflow:
            continue
        key = r.get("entity_name") or (r.get("memo") or "Other")
        agg[key] += amt
    items = [{"name": k, "amount": round(v, 2)} for k, v in agg.items()]
    items.sort(key=lambda x: abs(x["amount"]), reverse=True)
    return items


def cash_in_project_split(proj_rows: list[dict], cash_in_total: float) -> dict:
    """Deterministic split of the day's collections: Production itemized by project (#/name),
    everything else (Service-class or no GL project) summed as Service. Service is the residual
    against the authoritative AR inflow total, so Production + Service always ties to cash_in."""
    prod: dict[tuple, float] = defaultdict(float)
    attributed = 0.0
    for r in proj_rows:
        amt = round(float(r.get("applied_amt") or 0), 2)
        attributed += amt
        num = r.get("project_num")
        if r.get("class_id") == CLASS_PRODUCTION and num:
            prod[(num, r.get("project_name") or num)] += amt
    production = [{"project_num": k[0], "project_name": k[1], "amount": round(v, 2)}
                  for k, v in prod.items()]
    production.sort(key=lambda x: abs(x["amount"]), reverse=True)
    prod_sum = round(sum(p["amount"] for p in production), 2)
    return {"production": production,
            "service": round(cash_in_total - prod_sum, 2),   # residual -> everything non-Production
            "attributed_total": round(attributed, 2)}


def ap_paid_project_split(ap_rows: list[dict], ap_total: float) -> dict:
    """Split AP cash out into job-tied sections, each itemized by job -- and, within each job, by
    vendor. Production-class jobs and non-Production (Service) jobs get their own sections; only
    genuinely non-project spend (overhead / SG&A) is 'Other', computed as the residual against the
    AP-disbursements total so Production + Service + Other always ties. The per-vendor amounts
    within a job sum to that job's total (same rows, one extra group key)."""
    def accumulate(pred):
        by_job: dict[tuple, float] = defaultdict(float)
        by_vend: dict[tuple, dict] = defaultdict(lambda: defaultdict(float))
        for r in ap_rows:
            num = r.get("project_num")
            if num and pred(r):
                amt = round(float(r.get("amt") or 0), 2)
                key = (num, r.get("project_name") or num)
                by_job[key] += amt
                by_vend[key][r.get("vendor") or "(no vendor)"] += amt
        out = []
        for k, v in by_job.items():
            vendors = [{"vendor": vn, "amount": round(va, 2)} for vn, va in by_vend[k].items()]
            vendors.sort(key=lambda x: abs(x["amount"]), reverse=True)
            out.append({"project_num": k[0], "project_name": k[1],
                        "amount": round(v, 2), "vendors": vendors})
        out.sort(key=lambda x: abs(x["amount"]), reverse=True)
        return out
    production = accumulate(lambda r: r.get("class_id") == CLASS_PRODUCTION)
    service    = accumulate(lambda r: r.get("class_id") != CLASS_PRODUCTION)
    attributed = round(sum(round(float(r.get("amt") or 0), 2) for r in ap_rows), 2)
    prod_sum = round(sum(p["amount"] for p in production), 2)
    svc_sum  = round(sum(s["amount"] for s in service), 2)
    return {"production": production,
            "service": service,
            "other": round(ap_total - prod_sum - svc_sum, 2),
            "attributed_total": attributed}


def build_summary(report: dt.date, prior_date: dt.date, prior_total: float,
                  prior_source: str, bal_today: list[dict], movements: list[dict],
                  ar_total: float, ap_total: float,
                  proj_rows: list[dict] | None = None, unpaid: dict | None = None,
                  ap_rows: list[dict] | None = None,
                  journal_details: dict | None = None) -> dict:
    by_acct = {r["account_id"]: float(r["balance"]) for r in bal_today}
    names = {r["account_id"]: r["fullname"] for r in bal_today}
    total_today = sum(by_acct.values())
    net_change = total_today - prior_total

    buckets: dict[str, float] = defaultdict(float)
    movement_net = 0.0
    for r in movements:
        amt = signed(r["amount"])
        movement_net += amt
        buckets[classify(r)] += amt

    bucket_list = sorted(({"bucket": b, "net": round(v, 2)} for b, v in buckets.items()),
                         key=lambda x: abs(x["net"]), reverse=True)

    recon_diff = round((prior_total + movement_net) - total_today, 2)
    recon_ok = abs(recon_diff) <= EPSILON

    drivers = sorted((r for r in movements if classify(r) != "Internal transfer"),
                     key=lambda r: abs(signed(r["amount"])), reverse=True)
    top_movers = [{"tranid": r.get("tranid"), "date": r.get("trandate"),
                   "type": r.get("type_label"), "entity": r.get("entity_name"),
                   "account": r.get("account_name"), "memo": r.get("memo"),
                   "amount": round(signed(r["amount"]), 2), "bucket": classify(r)}
                  for r in drivers[:5]]

    cash_in_items = _itemize(movements, inflow=True)
    cash_out_all = _itemize(movements, inflow=False)
    cash_in_total = round(sum(i["amount"] for i in cash_in_items), 2)
    proj_split = cash_in_project_split(proj_rows or [], cash_in_total)
    ap_out_total = round(abs(buckets.get("AP disbursements", 0.0)), 2)
    ap_split = ap_paid_project_split(ap_rows or [], ap_out_total)
    cash_out_total = round(sum(i["amount"] for i in cash_out_all), 2)
    cash_out_items = cash_out_all[:TOP_CEO_ITEMS]
    others = cash_out_all[TOP_CEO_ITEMS:]
    cash_out_other = round(sum(i["amount"] for i in others), 2) if others else 0.0

    flags = []
    if not recon_ok:
        flags.append(f"RECONCILIATION MISMATCH: logged prior {prior_total:,.2f} + movements "
                     f"{movement_net:,.2f} = {prior_total + movement_net:,.2f}, but computed "
                     f"today {total_today:,.2f} (diff {recon_diff:,.2f}). Treat as provisional.")
    if prior_source == "bootstrap":
        flags.append(f"First run / no prior reported balance on file -- prior balance was "
                     f"reconstructed as of {prior_date} excluding entries back-posted "
                     f"afterward; any such back-posts appear in today's movements. Relies on "
                     f"createddate reflecting when entries hit the books.")
    for b in bucket_list:
        if abs(b["net"]) >= THRESHOLD and b["bucket"] != "Internal transfer":
            flags.append(f"Bucket over threshold: {b['bucket']} net {b['net']:,.2f}")
    for m in top_movers:
        if abs(m["amount"]) >= THRESHOLD:
            flags.append(f"Large single item: {m['type']} {m['amount']:,.2f} "
                         f"({m['entity'] or m['memo'] or m['tranid']})")
    other_net = buckets.get("Other / unclassified", 0.0)
    if abs(other_net) > EPSILON:
        flags.append(f"Unclassified cash movement of {other_net:,.2f} -- needs a sec.7 rule.")
    for aid, bal in by_acct.items():
        if bal < 0:
            flags.append(f"Negative balance: {names.get(aid, aid)} at {bal:,.2f}")
    if proj_rows is not None and abs(proj_split["attributed_total"] - cash_in_total) > max(EPSILON, 0.005 * abs(cash_in_total)):
        flags.append(f"Cash-in-by-project attribution {proj_split['attributed_total']:,.2f} differs "
                     f"from AR inflow {cash_in_total:,.2f}; residual folded into Service -- check for "
                     f"payments not applied to invoices, or non-customer inflows.")
    if ap_rows is not None and abs(ap_split["attributed_total"] - ap_out_total) > max(EPSILON, 0.005 * abs(ap_out_total)):
        flags.append(f"AP-by-project attribution {ap_split['attributed_total']:,.2f} differs from AP "
                     f"disbursements {ap_out_total:,.2f}; residual folded into Other -- check for "
                     f"disbursement types beyond VendPymt/Check or unusual postings.")

    return {
        "report_date": report.isoformat(),
        "prior_date": prior_date.isoformat(),
        "prior_source": prior_source,
        "currency_note": "positive = cash in, negative = cash out",
        "total_cash_today": round(total_today, 2),
        "total_cash_prior": round(prior_total, 2),
        "net_change": round(net_change, 2),
        "ar_total": round(ar_total, 2),
        "ap_total": round(ap_total, 2),
        # Aggregate only in the JSON payload; the full bill list rides in the overdue workbook.
        "unpaid_bills": ({k: unpaid[k] for k in ("basis", "asof", "count", "total") if k in unpaid}
                         if unpaid else None),
        "journal_details": journal_details or {},   # {journal_no: {offsets:[...], suggested_purpose?}}
        "account_balances": [{"account": names.get(aid, aid), "balance": round(by_acct[aid], 2),
                              "type": next((r["accttype"] for r in bal_today
                                            if r["account_id"] == aid), "")}
                             for aid in sorted(by_acct, key=lambda a: str(names.get(a, a)))],
        "by_bucket": bucket_list,
        "top_movers": top_movers,
        "cash_in": {"total": cash_in_total,
                    "production": proj_split["production"],
                    "service": proj_split["service"]},
        "ap_out": {"total": ap_out_total,
                   "production": ap_split["production"],
                   "service": ap_split["service"],
                   "other": ap_split["other"]},
        "cash_out": {"total": cash_out_total, "items": cash_out_items,
                     "all_others": cash_out_other},
        "reconciles": recon_ok,
        "flags": flags,
    }


# --------------------------------------------------------------------------------------
# LLM narrative -- CFO section + CEO text-message block (one call, two delimited parts)
# --------------------------------------------------------------------------------------
CEO_STYLE_EXAMPLE = """\
6/30

Cash: $4.1m
AR: $5.7m
AP: $2.6m
Overdue bills: $70k (23)

Cash in: $64k
Production
  Stanly County EOC (1471CLT) - $58k
Service - $6k

Cash out: $61k
Wild Edge Woodcraft - $53k
MBFS - $2k
All others $6k
"""


def write_sections(summary: dict) -> tuple[str, str]:
    from anthropic import Anthropic

    kb = ""
    if os.path.exists(KB_PATH):
        with open(KB_PATH, encoding="utf-8") as f:
            kb = f.read()

    system = (
        "You produce a daily cash report from FINISHED numbers. Do not recompute, "
        "re-categorize, or sum -- the figures are authoritative. Output PLAIN TEXT only: no "
        "Markdown, no asterisks, no pipe tables (the email is plain text). Return EXACTLY two "
        "sections separated by the delimiters shown, nothing before or after:\n"
        "<<<CFO>>>\n"
        "A precise CFO brief. One-line headline, then the drivers (by_bucket and top_movers), "
        "then note collections (cash_in: the Production projects and the Service total), AP paid by "
        "job (ap_out: the Production jobs and the Service jobs -- each job in production[] and "
        "service[] has a per-vendor breakdown in [].vendors -- plus the Other/overhead total; call "
        "out notable job->vendor concentration, e.g. a single sub/supplier that is most of a job's "
        "spend), and overdue unpaid bills "
        "(unpaid_bills), then every item in 'flags'. Keep full dollar figures with cents. If "
        "'reconciles' is false, lead with that and call the numbers provisional.\n"
        "For any Journal in top_movers, look it up in 'journal_details' by its tranid and STATE ITS "
        "PURPOSE from the offset accounts rather than calling it unclassified: if "
        "'suggested_purpose' is present use it verbatim; otherwise infer from the offsets (e.g. a "
        "single liability account plus an Interest Expense line is a debt principal-and-interest "
        "payment -- name it 'Debt Payment - <liability account>'). Name the offset accounts and "
        "amounts. Do not invent a purpose when the offsets don't support one.\n"
        "<<<CEO>>>\n"
        "A short text-message-style summary matching the example's format and rounding "
        "(round to $k/$m, abbreviate vendor/customer names sensibly). Rules:\n"
        "- Header lines: Cash, AR, AP, then Overdue bills. unpaid_bills is an object "
        "{basis, count, total}; render 'Overdue bills: $Xk (N)' from total and count. If it is null, "
        "write 'Overdue bills: [tbd]'.\n"
        "- Cash in: use cash_in.total for the headline amount. If cash_in.production is non-empty, "
        "add a 'Production' label then one line per project as 'ProjectName (project_num) - $Xk'; "
        "omit the label and lines if it is empty. Then 'Service - $Xk' from cash_in.service.\n"
        "- Cash out: itemize cash_out.items, then 'All others $Xk' for cash_out.all_others.\n"
        f"Format example:\n{CEO_STYLE_EXAMPLE}"
    )
    user = (f"KNOWLEDGE BASE:\n{kb}\n\nTODAY'S COMPUTED SUMMARY (JSON):\n"
            f"{json.dumps(summary, indent=2, default=str)}")

    client = Anthropic()
    resp = client.messages.create(model=ANTHROPIC_MODEL, max_tokens=2000,
                                  system=system, messages=[{"role": "user", "content": user}])
    text = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
    cfo, ceo = text, ""
    if "<<<CEO>>>" in text:
        cfo, ceo = text.split("<<<CEO>>>", 1)
    cfo = cfo.replace("<<<CFO>>>", "").strip()
    return cfo, ceo.strip()


# --------------------------------------------------------------------------------------
# Delivery
# --------------------------------------------------------------------------------------
def movements_csv(movements: list[dict]) -> str:
    buf = io.StringIO()
    cols = ["tranid", "trandate", "createddate", "type_label", "account_name",
            "accttype", "entity_name", "memo", "amount"]
    w = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    w.writeheader()
    for r in movements:
        w.writerow(r)
    return buf.getvalue()


def payments_workbook(proj_rows: list[dict], ap_rows: list[dict] | None = None) -> bytes:
    """Excel of the day's cash by project. AR (customer payments): AR detail (payment x invoice
    with customer + job), AR by job, AR by customer. AP (vendor cash out): AP detail (payment/check
    with vendor + job) and AP by job. Values are computed here (deterministic); no formulas, since
    the CI runner has no spreadsheet engine to recalculate them."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    MONEY = "#,##0.00"
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E78")
    BASE = Font(name="Arial")
    BOLD = Font(name="Arial", bold=True)
    CENTER = Alignment(horizontal="center")
    label = {CLASS_PRODUCTION: "Production", CLASS_SERVICE: "Service"}
    ap_rows = ap_rows or []

    def style_sheet(ws, ncols, widths, money_cols):
        total_row = ws.max_row
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c)
            cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, CENTER
        for r in range(2, ws.max_row + 1):
            for c in range(1, ncols + 1):
                ws.cell(r, c).font = BOLD if r == total_row else BASE
            for c in money_cols:
                ws.cell(r, c).number_format = MONEY
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = wdt
        ws.freeze_panes = "A2"

    wb = Workbook()

    # ---- AR: customer payments received ----
    def ar_amt(r): return round(float(r.get("applied_amt") or 0), 2)
    ar = sorted(proj_rows, key=lambda r: (str(r.get("customer") or ""),
                                          str(r.get("payment_no") or ""),
                                          str(r.get("invoice_no") or "")))
    ar_total = round(sum(ar_amt(r) for r in ar), 2)

    ws = wb.active
    ws.title = "AR detail"
    ws.append(["Payment #", "Date", "Customer", "Invoice #", "Job #", "Job name", "Class", "Amount"])
    for r in ar:
        ws.append([r.get("payment_no"), str(r.get("payment_date") or "")[:10], r.get("customer"),
                   r.get("invoice_no"), r.get("project_num") or "", r.get("project_name") or "",
                   label.get(r.get("class_id"), ""), ar_amt(r)])
    ws.append(["", "", "", "", "", "", "Total", ar_total])
    style_sheet(ws, 8, [15, 12, 34, 16, 12, 30, 12, 16], (8,))

    prod: dict[tuple, float] = defaultdict(float)
    service = 0.0
    for r in ar:
        if r.get("class_id") == CLASS_PRODUCTION and r.get("project_num"):
            prod[(r["project_num"], r.get("project_name") or r["project_num"])] += ar_amt(r)
        else:
            service += ar_amt(r)
    ws2 = wb.create_sheet("AR by job")
    ws2.append(["Job #", "Job name", "Amount"])
    for (num, name), v in sorted(prod.items(), key=lambda kv: kv[1], reverse=True):
        ws2.append([num, name, round(v, 2)])
    if round(service, 2):
        ws2.append(["", "Service (all non-Production)", round(service, 2)])
    ws2.append(["", "Total", ar_total])
    style_sheet(ws2, 3, [14, 40, 16], (3,))

    bycust: dict[str, float] = defaultdict(float)
    for r in ar:
        bycust[r.get("customer") or "(unknown)"] += ar_amt(r)
    ws3 = wb.create_sheet("AR by customer")
    ws3.append(["Customer", "Amount"])
    for name, v in sorted(bycust.items(), key=lambda kv: kv[1], reverse=True):
        ws3.append([name, round(v, 2)])
    ws3.append(["Total", ar_total])
    style_sheet(ws3, 2, [40, 16], (2,))

    # ---- AP: vendor cash paid out ----
    def ap_amt(r): return round(float(r.get("amt") or 0), 2)
    ap = sorted(ap_rows, key=lambda r: (str(r.get("vendor") or ""), str(r.get("payment_no") or "")))
    ap_total = round(sum(ap_amt(r) for r in ap), 2)

    ws4 = wb.create_sheet("AP detail")
    ws4.append(["Payment #", "Date", "Vendor", "Type", "Bill / account",
                "Job #", "Job name", "Class", "Amount"])
    for r in ap:
        ws4.append([r.get("payment_no"), str(r.get("payment_date") or "")[:10], r.get("vendor"),
                    r.get("source"), r.get("reference"), r.get("project_num") or "",
                    r.get("project_name") or "", label.get(r.get("class_id"), ""), ap_amt(r)])
    ws4.append(["", "", "", "", "", "", "", "Total", ap_total])
    style_sheet(ws4, 9, [15, 12, 30, 9, 22, 12, 28, 12, 16], (9,))

    # AP grouped by job, split into Production and Service sections; overhead stays in Other.
    def job_key(r): return (r["project_num"], r.get("project_name") or r["project_num"])
    prod_jobs: dict[tuple, float] = defaultdict(float)
    svc_jobs: dict[tuple, float] = defaultdict(float)
    other = 0.0
    for r in ap:
        if not r.get("project_num"):
            other += ap_amt(r)
        elif r.get("class_id") == CLASS_PRODUCTION:
            prod_jobs[job_key(r)] += ap_amt(r)
        else:
            svc_jobs[job_key(r)] += ap_amt(r)
    ws5 = wb.create_sheet("AP by job")
    ws5.append(["Class", "Job #", "Job name", "Amount"])
    for cls, jobs in (("Production", prod_jobs), ("Service", svc_jobs)):
        for (num, name), v in sorted(jobs.items(), key=lambda kv: kv[1], reverse=True):
            ws5.append([cls, num, name, round(v, 2)])
    if round(other, 2):
        ws5.append(["", "", "Other / non-project", round(other, 2)])
    ws5.append(["", "", "Total", ap_total])
    style_sheet(ws5, 4, [12, 14, 40, 16], (4,))

    # AP rolled up by job, then vendor -- Production and Service sections; overhead in Other.
    def jv_block(pred):
        jv: dict[tuple, float] = defaultdict(float)
        for r in ap:
            if r.get("project_num") and pred(r):
                jv[(r["project_num"], r.get("project_name") or r["project_num"],
                    r.get("vendor") or "(no vendor)")] += ap_amt(r)
        jt: dict[tuple, float] = defaultdict(float)
        for (num, name, _v), amt in jv.items():
            jt[(num, name)] += amt
        return jv, jt
    ws6 = wb.create_sheet("AP by job x vendor")
    ws6.append(["Class", "Job #", "Job name", "Vendor", "Amount"])
    for cls, pred in (("Production", lambda r: r.get("class_id") == CLASS_PRODUCTION),
                      ("Service", lambda r: r.get("class_id") != CLASS_PRODUCTION)):
        jv, jt = jv_block(pred)
        for (num, name) in sorted(jt, key=lambda k: jt[k], reverse=True):
            vend = sorted([(v, a) for (n, nm, v), a in jv.items() if (n, nm) == (num, name)],
                          key=lambda x: x[1], reverse=True)
            for v, a in vend:
                ws6.append([cls, num, name, v, round(a, 2)])
    if round(other, 2):
        ws6.append(["", "", "Other / non-project", "", round(other, 2)])
    ws6.append(["", "", "", "Total", ap_total])
    style_sheet(ws6, 5, [12, 14, 34, 30, 16], (5,))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def overdue_bills_workbook(unpaid: dict) -> bytes:
    """One-sheet Excel of every open vendor bill due on or before the report date, so the AP total
    can be reconciled line by line. Most-overdue first; values only (CI has no recalc engine)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    MONEY = "#,##0.00"
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1F4E78")
    BASE = Font(name="Arial")
    BOLD = Font(name="Arial", bold=True)
    CENTER = Alignment(horizontal="center")

    bills = sorted(unpaid.get("bills", []),
                   key=lambda b: (b.get("days_overdue") is None, -(b.get("days_overdue") or 0)))
    total = round(sum(float(b.get("unpaid") or 0) for b in bills), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Overdue bills"
    ws.append(["Bill #", "Vendor", "Bill date", "Due date", "Days overdue",
               "Bill amount", "Unpaid balance", "Status"])
    for b in bills:
        ws.append([b.get("bill_no") or "", b.get("vendor") or "",
                   str(b.get("trandate") or "")[:10], str(b.get("duedate") or "")[:10],
                   b.get("days_overdue"), round(float(b.get("bill_amount") or 0), 2),
                   round(float(b.get("unpaid") or 0), 2), b.get("status") or ""])
    ws.append(["", "", "", "", "", "", total, ""])
    total_row = ws.max_row
    for c in range(1, 9):
        cell = ws.cell(1, c)
        cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, CENTER
    for r in range(2, total_row + 1):
        for c in range(1, 9):
            ws.cell(r, c).font = BOLD if r == total_row else BASE
        ws.cell(r, 6).number_format = MONEY
        ws.cell(r, 7).number_format = MONEY
    ws.cell(total_row, 5).value = "Total"
    ws.cell(total_row, 5).font = BOLD
    for i, wdt in enumerate([16, 34, 12, 12, 13, 15, 16, 9], 1):
        ws.column_dimensions[chr(64 + i)].width = wdt
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def cfo_detail_block(summary: dict) -> str:
    """Deterministic long-form, full values with cents -- the audit view."""
    lines = [f"Cash today ({summary['report_date']}): {summary['total_cash_today']:,.2f}"]
    for a in summary["account_balances"]:
        lines.append(f"    {a['account']:<26} {a['balance']:>16,.2f}  [{a['type']}]")
    lines.append(f"Prior reported ({summary['prior_date']}): {summary['total_cash_prior']:,.2f}"
                 f"   (source: {summary['prior_source']})")
    lines.append(f"Net change: {summary['net_change']:,.2f}")
    lines.append(f"AR total: {summary['ar_total']:,.2f}    AP total: {summary['ap_total']:,.2f}")
    lines.append("")
    lines.append("By bucket:")
    for b in summary["by_bucket"]:
        lines.append(f"    {b['bucket']:<22} {b['net']:>16,.2f}")
    lines.append("")
    ci = summary["cash_in"]
    lines.append(f"Cash in by project: {ci['total']:,.2f}")
    if ci.get("production"):
        lines.append("    Production:")
        for p in ci["production"]:
            lines.append(f"        {p['project_num']:<10} {(p['project_name'] or ''):<28} "
                         f"{p['amount']:>14,.2f}")
    lines.append(f"    {'Service:':<39}{ci['service']:>14,.2f}")
    ap = summary.get("ap_out")
    if ap:
        lines.append("")
        lines.append(f"AP paid by project: {ap['total']:,.2f}")
        for section, heading in (("production", "Production (by job, then vendor):"),
                                 ("service", "Service (by job, then vendor):")):
            if ap.get(section):
                lines.append(f"    {heading}")
                for p in ap[section]:
                    lines.append(f"        {p['project_num']:<10} {(p['project_name'] or ''):<28} "
                                 f"{p['amount']:>14,.2f}")
                    for v in p.get("vendors", []):
                        lines.append(f"            {(v['vendor'] or ''):<32} {v['amount']:>14,.2f}")
        lines.append(f"    {'Other / non-project:':<39}{ap['other']:>14,.2f}")
    ub = summary.get("unpaid_bills")
    if ub:
        lines.append("")
        lines.append(f"Overdue unpaid bills (due on or before {summary['report_date']}): "
                     f"{ub['total']:,.2f}  ({ub['count']} bills) -- see the overdue-bills Excel")
    jd = summary.get("journal_details") or {}
    if jd:
        lines.append("")
        lines.append("Journal purposes (offsets, net of clearing):")
        for jn, d in jd.items():
            label = d.get("suggested_purpose")
            lines.append(f"    {jn}" + (f"  -- {label}" if label else ""))
            for o in d.get("offsets", []):
                lines.append(f"        {(o['account'] or ''):<34} {o['amount']:>14,.2f}  "
                             f"[{o['accttype']}]")
    lines.append("")
    lines.append(f"Reconciles (roll-forward): {summary['reconciles']}")
    return "\n".join(lines)


def send_email(summary: dict, cfo_narrative: str, ceo_block: str, movements: list[dict],
               proj_rows: list[dict] | None = None, ap_rows: list[dict] | None = None,
               unpaid: dict | None = None) -> None:
    msg = EmailMessage()
    msg["Subject"] = f"Cash snap -- {summary['report_date']}"
    msg["From"], msg["To"] = MAIL_FROM, MAIL_TO
    body = (
        "===== CFO VIEW =====\n\n"
        f"{cfo_narrative}\n\n"
        "----- detail (full values) -----\n"
        f"{cfo_detail_block(summary)}\n\n\n"
        "===== CEO OUTPUT =====\n\n"
        f"{ceo_block}\n\n"
        "(full numbers in the attached JSON; cash line detail in the CSV; cash in/out by "
        "invoice/job/customer/vendor in the payments Excel; every overdue bill line by line in "
        "the overdue-bills Excel)\n"
    )
    msg.set_content(body)
    msg.add_attachment(json.dumps(summary, indent=2, default=str).encode(),
                       maintype="application", subtype="json",
                       filename=f"cash_snap_{summary['report_date']}.json")
    msg.add_attachment(movements_csv(movements).encode(), maintype="text", subtype="csv",
                       filename=f"cash_movements_{summary['report_date']}.csv")
    if proj_rows is not None or ap_rows is not None:
        msg.add_attachment(
            payments_workbook(proj_rows or [], ap_rows or []), maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"payments_by_project_{summary['report_date']}.xlsx")
    if unpaid and unpaid.get("bills"):
        msg.add_attachment(
            overdue_bills_workbook(unpaid), maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"overdue_bills_{summary['report_date']}.xlsx")
    _smtp_send(msg)


def send_alert(error: str) -> None:
    msg = EmailMessage()
    msg["Subject"] = "Cash snap FAILED"
    msg["From"], msg["To"] = MAIL_FROM, MAIL_TO
    msg.set_content(f"The cash snap pipeline failed and produced no report.\n\n{error}")
    try:
        _smtp_send(msg)
    except Exception:
        pass


def _smtp_send(msg: EmailMessage) -> None:
    if SMTP_PORT == 465:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30) as s:
            s.login(SMTP_USER, SMTP_PASS)
            s.send_message(msg)
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.send_message(msg)


# --------------------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------------------
def main() -> int:
    try:
        report = resolve_report_date()
        bal_today = balances_as_of(report, (), extra_account_ids=CASH_ACCOUNT_IDS)
        if not bal_today:
            raise RuntimeError("No cash-account balances returned -- check role permissions and "
                               "the CASH_ACCOUNT_IDS allowlist (default 223 First Bank, 122 UF).")
        total_today = sum(float(r["balance"]) for r in bal_today)

        state = load_state()
        if state and state.get("report_date"):
            prior_date = dt.date.fromisoformat(state["report_date"])
            prior_total = float(state["total_cash"])
            prior_source = "logged"
            include_created = True
        else:
            prior_date = previous_business_day(report)
            # Reconstruct the prior balance as it actually stood at prior_date's close --
            # EXCLUDING anything back-posted afterward (created after prior_date). Those
            # back-posts then surface in this run's movements via include_created=True, so the
            # first run is a true roll-forward instead of a naive recompute that would bury them.
            prior_total = total_balance(prior_date, (),
                                        created_on_or_before=prior_date,
                                        extra_account_ids=CASH_ACCOUNT_IDS)
            prior_source = "bootstrap"
            include_created = True

        movements = cash_movements(prior_date, report, include_created)
        ar_total = abs(total_balance(report, ("AcctRec",)))
        ap_total = abs(total_balance(report, ("AcctPay",)))
        proj_rows = cash_in_by_project(prior_date, report, include_created)
        unpaid = overdue_bills(report)
        ap_rows = ap_paid_by_project(prior_date, report, include_created)
        journal_details = journal_offsets(movements)

        summary = build_summary(report, prior_date, prior_total, prior_source,
                                bal_today, movements, ar_total, ap_total,
                                proj_rows=proj_rows, unpaid=unpaid, ap_rows=ap_rows,
                                journal_details=journal_details)
        cfo_narrative, ceo_block = write_sections(summary)
        send_email(summary, cfo_narrative, ceo_block, movements,
                   proj_rows=proj_rows, ap_rows=ap_rows, unpaid=unpaid)

        # Log only after a successful send; the workflow commits this file.
        save_state(report, total_today, bal_today)
        print(f"Cash snap sent for {report.isoformat()} "
              f"(reconciles={summary['reconciles']}, flags={len(summary['flags'])}).")
        return 0
    except Exception as exc:  # noqa: BLE001
        send_alert(f"{type(exc).__name__}: {exc}")
        print(f"FAILED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
